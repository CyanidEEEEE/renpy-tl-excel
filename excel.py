import os
import re
import openpyxl
from openpyxl.utils import get_column_letter
import multiprocessing
import time
from multiprocessing import Manager, Pool
from tqdm import tqdm
import concurrent.futures

# 优化后的正则表达式
STRING_PATTERN = re.compile(
    r"(#\s+game/(.*?:\d+))?\s*translate\s+\w+\s+strings:\s*(?:\s*#.*\n)?\s*(?:(#\s+.*?(/.*?:\d+))?\s*old\s*\"(.*?)\"\s*\n\s*new\s*\"(.*?)\"\s*)+"
)
INDIVIDUAL_STRING_PATTERN = re.compile(
    r"(#\s+(.*?(/.*?:\d+)))?\s*old\s*\"(.*?)\"\s*\n\s*new\s*\"(.*?)\""
)
DIALOGUE_PATTERN = re.compile(
    r"(#\s+game/(.*?:\d+))?\s*translate\s+\w+\s+(\w*):?\s*(?:\s*#.*?\n)?\s*(#\s*((\w+)?)\s*\"(.*?)\")(?:\s*[\r\n]+\s*(?:\w+)?\s*\"(.*?)\")?(?=\s*\n(?:#|$|translate\s+\w+\s+strings))"
)

def extract_translation_data(rpy_file_path: str, language: str) -> list:
    """从 .rpy 文件中提取翻译数据，用于导出。"""
    data = []
    try:
        with open(rpy_file_path, 'r', encoding='utf-8') as f:
            content = f.read()

            # 匹配字符串翻译
            for match in STRING_PATTERN.finditer(content):
                location_prefix = match.group(2) if match.group(2) else ""
                for individual_match in INDIVIDUAL_STRING_PATTERN.finditer(match.group(0)):
                    # 使用 group(3) 来获取 location
                    location = individual_match.group(3).strip() if individual_match.group(3) else location_prefix
                    location = location.split('/')[-1] if location else ""
                    original = individual_match.group(4).strip()
                    translation = individual_match.group(5).strip()
                    data.append({
                        "Prefix": "strings",
                        "Original": original,
                        "Translation": translation,
                        "Location": location,
                        "Identifier": ""
                    })

            # 匹配对话翻译
            for match in DIALOGUE_PATTERN.finditer(content):
                location = match.group(2) if match.group(2) else ""
                location = location.split('/')[-1] if location else ""

                # 如果没有 location，则尝试从注释中提取
                if not location:
                    location_comment = match.group(1)
                    if location_comment:
                        location_match = re.search(r"game/(.*?:\d+)", location_comment)
                        if location_match:
                            location = location_match.group(1)

                identifier = match.group(3) if match.group(3) else ""

                prefix = match.group(6) if match.group(6) else ""
                original = match.group(7).strip() if match.group(7) else ""
                translation = match.group(8).strip() if match.group(8) else ""
                data.append({
                    "Prefix": prefix,
                    "Original": original,
                    "Translation": translation,
                    "Location": location,
                    "Identifier": identifier
                })

    except Exception as e:
        print(f"  [ERROR] 文件: {rpy_file_path}, 发生错误: {type(e).__name__} - {e}")
    return data

def process_rpy_file(args):
    """处理单个 .rpy 文件并返回数据，用于导出。"""
    rpy_file_path, language, shared_data = args
    data = extract_translation_data(rpy_file_path, language)
    shared_data.extend(data)

def export_to_excel(tl_folder_path: str, language: str, output_excel_file: str):
    """将翻译数据导出到 Excel 文件。"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # 修改列名，并添加一列空的占位列，第一行改为“特殊”
    sheet.append(["前缀", "原文", "译文", "特殊", "定位", "标识"])

    rpy_files = [os.path.join(root, file) for root, _, files in os.walk(tl_folder_path) for file in files if file.endswith(".rpy")]
    total_files = len(rpy_files)

    with Manager() as manager:
        shared_data = manager.list()
        pool_size = min(multiprocessing.cpu_count(), total_files)

        with multiprocessing.Pool(pool_size) as pool:
            with tqdm(total=total_files, desc="处理文件") as pbar:
                for _ in pool.imap_unordered(process_rpy_file, [(rpy_file, language, shared_data) for rpy_file in rpy_files]):
                    pbar.update()

        # 批量写入 Excel，将定位和标识写入到第五、六列
        data_to_write = []
        for item in shared_data:
            # 对导出excel的行进行调整，当prefix为空时，不填入narrator
            data_to_write.append([item["Prefix"], item["Original"], item["Translation"], "", item["Location"], item["Identifier"]])
        for row in tqdm(data_to_write, desc="写入 Excel"):
            sheet.append(row)

        # 设置列宽
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 25
        workbook.save(output_excel_file)

def conditional_patch(game_root: str, excel_file: str):
    """条件修补功能"""
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        print("--- 开始条件修补 ---")
        print(f"Excel 文件: {excel_file}")
        print(f"游戏根目录: {game_root}")

        # 检查 Excel 文件是否为空，如果为空则写入表头
        if sheet.max_row <= 1:
            # 写入表头
            sheet.append(["前缀", "原文", "译文", "条件", "定位", "标识"])

        # 构建翻译映射, 现在只需要映射对话
        translation_map = {}
        for row in sheet.iter_rows(min_row=2):
            prefix, original, translation, _, location, identifier = [
                cell.value for cell in row
            ]
            if prefix != "strings" and location is not None:
                key = (
                    prefix,
                    original,
                    location.split(":")[0],
                )  # key现在只基于原文和文件名
                translation_map[key] = translation

        # 预编译正则表达式
        if_pattern = re.compile(
            r"(?s)(^\s*(if|elif|else)\s(.*?):.*?(?=\n\s*[a-zA-Z#]|\Z))", re.MULTILINE
        )
        dialogue_pattern = re.compile(r"(\w+)\s*\"(.*?)\"")  # 匹配对话行

        # 用于存储需要修改的行
        rows_to_modify = []

        # 遍历 game 文件夹（除了 tl 子文件夹）
        for root, dirs, files in os.walk(os.path.join(game_root, "game")):
            if "tl" in dirs:
                dirs.remove("tl")  # 排除 tl 文件夹
            for file in files:
                if file.endswith(".rpy"):
                    rpy_file_path = os.path.join(root, file)
                    print(f"\n--- 处理文件: {rpy_file_path} ---")
                    try:
                        with open(rpy_file_path, "r", encoding="utf-8") as f:
                            content = f.read()

                        # 寻找 if/elif/else 结构
                        matches = if_pattern.finditer(content)

                        current_if_condition = None
                        for match in matches:
                            statement = match.group(0)
                            condition = match.group(3).strip()

                            # 计算 if 语句的行号
                            line_num = content.count("\n", 0, match.start()) + 1

                            if statement.lstrip().startswith("if"):
                                current_if_condition = condition
                            elif statement.lstrip().startswith("elif"):
                                pass  # elif 不改变 current_if_condition
                            elif statement.lstrip().startswith("else"):
                                condition = f"not ({current_if_condition})"

                            print(f"  匹配到条件语句:\n{statement}")
                            print(f"    语句起始行号: {line_num}")
                            print(f"    提取到条件: {condition}")

                            # 获取 if 语句块内的所有行
                            block_start = match.end()
                            next_if = content.find("if ", block_start)
                            next_elif = content.find("elif ", block_start)
                            next_else = content.find("else:", block_start)

                            if next_if == -1:
                                next_if = float("inf")
                            if next_elif == -1:
                                next_elif = float("inf")
                            if next_else == -1:
                                next_else = float("inf")

                            block_end = min(next_if, next_elif, next_else, len(content))

                            block_lines = content[block_start:block_end].splitlines()

                            for block_line in block_lines:
                                original = block_line.strip()

                                # 移除行尾注释
                                original = original.split("#")[0].strip()

                                if not original:  # 跳过空行
                                    continue

                                # 跳过特定类型的行,例如$"..."的
                                if original.startswith("$"):
                                    continue

                                # 判断是否为控制语句
                                if original.startswith(
                                    (
                                        "jump ",
                                        "scene ",
                                        "show ",
                                        "hide ",
                                        "with ",
                                        "play ",
                                        "stop ",
                                        "pause ",
                                        "call ",
                                        "return ",
                                    )
                                ):
                                    continue

                                # 判断是否为menu:
                                if original == "menu:":
                                    continue

                                # 尝试匹配对话行
                                dialogue_match = dialogue_pattern.match(original)

                                if dialogue_match:
                                    prefix = dialogue_match.group(1)
                                    original = dialogue_match.group(2)
                                else:
                                    prefix = ""
                                    print(f"  标记为无前缀")

                                # 移除首尾的引号
                                original = (
                                    original[1:-1]
                                    if original.startswith('"')
                                    and original.endswith('"')
                                    else original
                                )

                                # 构建键值
                                file_name = os.path.basename(rpy_file_path)
                                key = (
                                    prefix,
                                    original,
                                    file_name,
                                )  # key现在包含前缀，原文和文件名
                                print(
                                    f"    前缀: {prefix}, 原文: {original}, 文件名: {file_name}"
                                )

                                # 查找对应的 Excel 行
                                if key in translation_map:
                                    excel_row_index = None
                                    for row_index, row in enumerate(
                                        sheet.iter_rows(min_row=2), start=2
                                    ):
                                        # 检查键的每个部分是否匹配
                                        if (
                                            row[0].value == key[0]
                                            and row[1].value == key[1]
                                            and row[4].value.startswith(key[2] + ":")
                                        ):
                                            excel_row_index = row_index
                                            break

                                    if excel_row_index != -1:
                                        print(
                                            f"      在 Excel 中找到匹配行: {excel_row_index}"
                                        )
                                        # 添加到待修改行列表
                                        rows_to_modify.append((excel_row_index, condition))
                                    else:
                                        print(f"      未在 Excel 中找到匹配行")
                                else:
                                    print(f"      键值 {key} 未在翻译映射中找到")

                    except Exception as e:
                        print(
                            f"  [ERROR] 文件: {rpy_file_path}, 发生错误: {type(e).__name__} - {e}"
                        )

        # 批量修改 Excel 行
        for row_index, condition in tqdm(rows_to_modify, desc="更新 Excel"):
            condition_cell = sheet.cell(row=row_index, column=4)
            # 清除原有条件
            condition_cell.value = ""
            # 写入新条件
            condition_cell.value = condition

        workbook.save(excel_file)
        print("--- 条件修补完成 ---")

    except Exception as e:
        print(f"  [ERROR] 读取 Excel 文件 {excel_file} 失败: {type(e).__name__} - {e}")

if __name__ == '__main__':
    start_time = time.time()
    game_root = "." 
    tl_folder = os.path.join(game_root, "game", "tl")

    language_folder = input("请输入要操作的语言文件夹名称 (例如: chinese): ")
    language_path = os.path.join(tl_folder, language_folder)
    if not os.path.isdir(language_path):
        print("指定的语言文件夹不存在！")
        exit()

    operation = input("请选择操作类型 (1: 导入, 2: 导出, 3: 条件修补): ")

    excel_file = os.path.join(game_root, f"{language_folder}.xlsx")

    if operation == "1":
        print("导入功能尚未实现，请选择其他操作！")
        exit()
    elif operation == "2":
        if os.path.exists(excel_file):
            print(f"文件 {excel_file} 已存在，将会被覆盖。")
        export_to_excel(language_path, language_folder, excel_file)
        print(f"翻译数据已导出到 {excel_file}")
    elif operation == "3":
        conditional_patch(game_root, excel_file)
        print(f"条件修补已完成，结果已保存到 {excel_file}")
    else:
        print("无效的操作类型！")
    end_time = time.time()
    print(f"总耗时：{end_time - start_time} 秒")