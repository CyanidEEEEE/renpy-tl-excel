import os
import re
import openpyxl
from openpyxl.utils import get_column_letter
import multiprocessing
import time  # 确保这一行存在！
from multiprocessing import Manager, Pool
from tqdm import tqdm
import concurrent.futures

# 优化后的正则表达式 (保持不变)
DIALOGUE_PATTERN = re.compile(
    r"(\w+)?\s+\"(.*?)\"(?=\s*\n|$)|^\s*\"(.*?)\"(?=\s*\n|$)", re.MULTILINE
)
IF_BLOCK_PATTERN = re.compile(
    r"(?s)(^\s*(if|elif|else)\s(.*?):.*?(?=\n\s*[a-zA-Z#]|\Z))", re.MULTILINE
)


def build_translation_map(excel_file):
    """构建翻译映射和 Excel 行索引映射"""
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    translation_map = {}
    excel_row_index_map = {}  # 新增：Excel 行索引映射

    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2): # enumerate to get row_index
        prefix, original, translation, _, location, identifier = [
            cell.value for cell in row
        ]
        prefix = prefix if prefix else ""
        if prefix != "strings" and location is not None:
            key = (prefix, original, location.split(":")[0])
            translation_map[key] = translation

            # 构建 excel_row_index_map
            if key not in excel_row_index_map:
                excel_row_index_map[key] = [] # Initialize as a list to handle multiple matches
            excel_row_index_map[key].append(row_index) # Store the row_index

    return translation_map, sheet, excel_row_index_map # return excel_row_index_map as well


def process_rpy_file(rpy_file_path, translation_map, excel_row_index_map, rows_to_modify, sheet): # add excel_row_index_map as argument
    """处理单个 .rpy 文件"""
    print(f"\n--- 处理文件: {rpy_file_path} ---")
    try:
        with open(rpy_file_path, "r", encoding="utf-8") as f:
            content = f.read()
        lines = content.splitlines()
        if_blocks = IF_BLOCK_PATTERN.finditer(content)

        current_if_condition = None
        for if_block_match in if_blocks:
            statement = if_block_match.group(0)
            condition = if_block_match.group(3).strip()
            if_block_start_line_index = content.count("\n", 0, if_block_match.start())

            if statement.lstrip().startswith("if"):
                current_if_condition = condition
            elif statement.lstrip().startswith("elif"):
                pass
            elif statement.lstrip().startswith("else"):
                condition = f"not ({current_if_condition})"

            print(f"  匹配到条件语句:\n{statement}")
            print(f"    语句起始行号: {if_block_start_line_index + 1}")
            print(f"    提取到条件: {condition}")


            if_line_indentation = (
                len(lines[if_block_start_line_index])
                - len(lines[if_block_start_line_index].lstrip())
            )
            block_lines = []
            for line_index in range(
                if_block_start_line_index + 1, len(lines)
            ):
                current_line = lines[line_index]
                current_indentation = (
                    len(current_line) - len(current_line.lstrip())
                )
                if (
                    current_indentation > if_line_indentation
                ):
                    block_lines.append(current_line)
                else:
                    break

            for block_line in block_lines:
                original_line = block_line.strip()
                original = original_line

                original = original.split("#")[0].strip()

                if not original:
                    continue
                if original.startswith("$"):
                    continue
                if '"' not in original_line:
                    continue
                if original == "menu:":
                    continue

                dialogue_match = DIALOGUE_PATTERN.match(original)

                if dialogue_match:
                    prefix = (
                        dialogue_match.group(1)
                        if dialogue_match.group(1)
                        else ""
                    )
                    if not prefix:
                        prefix = ""
                        original = (
                            dialogue_match.group(3).strip()
                            if dialogue_match.group(3)
                            else ""
                        )
                    else:
                        original = (
                            dialogue_match.group(2).strip()
                            if dialogue_match.group(2)
                            else ""
                        )
                    print(f"  [DEBUG] Dialogue Match: Prefix='{prefix}', Original='{original}'")

                    file_name = os.path.basename(rpy_file_path)
                    key = (prefix, original, file_name)
                    print(
                        f"    前缀: {prefix}, 原文: {original}, 文件名: {file_name}"
                    )

                    if key in translation_map:
                        print(f"    [DEBUG] Key '{key}' found in translation_map")
                        if key in excel_row_index_map: # Directly use pre-indexed row indices
                            excel_row_indices = excel_row_index_map[key]
                            print(f"    [DEBUG] Found excel_row_indices from index: {excel_row_indices}")
                            if len(excel_row_indices) > 1: # 检查是否匹配到多个 Excel 行
                                print(f"    [DEBUG] 匹配到多个 Excel 行，标记为 'repeat'")
                                for excel_row_index in excel_row_indices:
                                    rows_to_modify.append(
                                        (excel_row_index, "repeat") # 直接填入 "repeat"
                                    )
                                    print(f"    [DEBUG] Added to rows_to_modify (repeat): row_index={excel_row_index}, condition='repeat'")
                            else: # 只有一个匹配行，按原逻辑处理
                                for excel_row_index in excel_row_indices:
                                    rows_to_modify.append(
                                        (excel_row_index, condition)
                                    )
                                    print(f"    [DEBUG] Added to rows_to_modify: row_index={excel_row_index}, condition='{condition}'")
                        else: # 这部分应该不会发生，除非索引构建有问题
                             print(f"      [ERROR] Key '{key}' not found in excel_row_index_map, but found in translation_map. Indexing issue?")
                    else:
                        print(f"      键值 {key} 未在翻译映射中找到")
                else:
                    prefix = ""
                    print(f"  标记为无前缀")

    except Exception as e:
        print(f"  [ERROR] 文件: {rpy_file_path}, 发生错误: {type(e).__name__} - {e}")


def update_excel_conditions(excel_file, rows_to_modify):
    """更新 Excel 文件中的条件列"""
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active  # 假设始终操作第一个 sheet，如果需要根据 sheet 名操作，需要修改
    for row_index, condition in tqdm(rows_to_modify, desc="更新 Excel"):
        condition_cell = sheet.cell(row=row_index, column=4)
        condition_cell.value = ""
        condition_cell.value = condition
    workbook.save(excel_file)


def conditional_patch_parallel(game_root: str, excel_file: str):
    """并行条件修补功能"""
    try:
        print("--- 开始条件修补 (并行) ---")
        print(f"Excel 文件: {excel_file}")
        print(f"游戏根目录: {game_root}")

        translation_map, sheet, excel_row_index_map = build_translation_map(excel_file) # Get excel_row_index_map here

        rows_to_modify = Manager().list()
        game_folder_path = os.path.join(game_root, "game")
        rpy_files = []
        for root, dirs, files in os.walk(game_folder_path):
            if "tl" in dirs:
                dirs.remove("tl")
            for file in files:
                if file.endswith(".rpy"):
                    rpy_files.append(os.path.join(root, file))

        cpu_count = multiprocessing.cpu_count()
        print(f"可用 CPU 核心数: {cpu_count}")
        pool = Pool(processes=cpu_count)
        tasks = [(file, translation_map, excel_row_index_map, rows_to_modify, sheet) for file in rpy_files] # pass excel_row_index_map to tasks

        with tqdm(total=len(rpy_files), desc="并行处理文件") as pbar:
            for _ in pool.imap_unordered(process_rpy_file_wrapper, tasks):
                pbar.update()
        pool.close()
        pool.join()

        update_excel_conditions(excel_file, list(rows_to_modify))

        print("--- 条件修补完成 (并行) ---")

    except Exception as e:
        print(f"  [ERROR] 并行处理失败: {type(e).__name__} - {e}")


def process_rpy_file_wrapper(task_args):
    """包装 process_rpy_file 以适应 imap_unordered"""
    file, translation_map, excel_row_index_map, rows_to_modify, sheet = task_args # unpack excel_row_index_map here
    return process_rpy_file(file, translation_map, excel_row_index_map, rows_to_modify, sheet) # pass excel_row_index_map


if __name__ == "__main__":
    start_time = time.time()
    game_root = "."
    tl_folder = os.path.join(game_root, "game", "tl")

    language_folder = input("请输入要操作的语言文件夹名称 (例如: chinese): ")
    language_path = os.path.join(tl_folder, language_folder)
    if not os.path.isdir(language_path):
        print("指定的语言文件夹不存在！")
        exit()

    excel_file = os.path.join(game_root, f"{language_folder}.xlsx")

    conditional_patch_parallel(game_root, excel_file)
    print(f"条件修补已完成，结果已保存到 {excel_file}")

    end_time = time.time()
    print(f"总耗时：{end_time - start_time} 秒")