import os
import re
import openpyxl
from openpyxl.utils import get_column_letter
import multiprocessing
import time
from multiprocessing import Manager, Pool
from tqdm import tqdm
import concurrent.futures

# 优化后的正则表达式
STRING_PATTERN = re.compile(
    r"(#\s+game/(.*?:\d+))?\s*translate\s+\w+\s+strings:\s*(?:\s*#.*\n)?\s*(?:(#\s+.*?(/.*?:\d+))?\s*old\s*\"(.*?)\"\s*\n\s*new\s*\"(.*?)\"\s*)+"
)
INDIVIDUAL_STRING_PATTERN = re.compile(
    r"(#\s+(.*?(/.*?:\d+)))?\s*old\s*\"(.*?)\"\s*\n\s*new\s*\"(.*?)\""
)
DIALOGUE_PATTERN = re.compile(
    r"(#\s+game/(.*?:\d+))?\s*translate\s+\w+\s+(\w*):?\s*(?:\s*#.*?\n)?\s*(#\s*((\w+)?)\s*\"(.*?)\")(?:\s*[\r\n]+\s*(?:\w+)?\s*\"(.*?)\")?(?=\s*\n(?:#|$|translate\s+\w+\s+strings))"
)

def extract_translation_data(rpy_file_path: str, language: str) -> list:
    """从 .rpy 文件中提取翻译数据，用于导出。"""
    data = []
    try:
        with open(rpy_file_path, 'r', encoding='utf-8') as f:
            content = f.read()

            # 匹配字符串翻译
            for match in STRING_PATTERN.finditer(content):
                location_prefix = match.group(2) if match.group(2) else ""
                for individual_match in INDIVIDUAL_STRING_PATTERN.finditer(match.group(0)):
                    # 使用 group(3) 来获取 location
                    location = individual_match.group(3).strip() if individual_match.group(3) else location_prefix
                    location = location.split('/')[-1] if location else ""
                    original = individual_match.group(4).strip()
                    translation = individual_match.group(5).strip()
                    data.append({
                        "Prefix": "strings",
                        "Original": original,
                        "Translation": translation,
                        "Location": location,
                        "Identifier": ""
                    })

            # 匹配对话翻译
            for match in DIALOGUE_PATTERN.finditer(content):
                location = match.group(2) if match.group(2) else ""
                location = location.split('/')[-1] if location else ""

                # 如果没有 location，则尝试从注释中提取
                if not location:
                    location_comment = match.group(1)
                    if location_comment:
                        location_match = re.search(r"game/(.*?:\d+)", location_comment)
                        if location_match:
                            location = location_match.group(1)

                identifier = match.group(3) if match.group(3) else ""

                prefix = match.group(6) if match.group(6) else ""
                original = match.group(7).strip() if match.group(7) else ""
                translation = match.group(8).strip() if match.group(8) else ""
                data.append({
                    "Prefix": prefix,
                    "Original": original,
                    "Translation": translation,
                    "Location": location,
                    "Identifier": identifier
                })

    except Exception as e:
        print(f"  [ERROR] 文件: {rpy_file_path}, 发生错误: {type(e).__name__} - {e}")
    return data

def process_rpy_file(args):
    """处理单个 .rpy 文件并返回数据，用于导出。"""
    rpy_file_path, language, shared_data = args
    data = extract_translation_data(rpy_file_path, language)
    shared_data.extend(data)

def export_to_excel(tl_folder_path: str, language: str, output_excel_file: str):
    """将翻译数据导出到 Excel 文件。"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # 修改列名，并添加一列空的占位列，第一行改为“特殊”
    sheet.append(["前缀", "原文", "译文", "特殊", "定位", "标识"])

    rpy_files = [os.path.join(root, file) for root, _, files in os.walk(tl_folder_path) for file in files if file.endswith(".rpy")]
    total_files = len(rpy_files)

    with Manager() as manager:
        shared_data = manager.list()
        pool_size = min(multiprocessing.cpu_count(), total_files)

        with multiprocessing.Pool(pool_size) as pool:
            with tqdm(total=total_files, desc="处理文件") as pbar:
                for _ in pool.imap_unordered(process_rpy_file, [(rpy_file, language, shared_data) for rpy_file in rpy_files]):
                    pbar.update()

        # 批量写入 Excel，将定位和标识写入到第五、六列
        data_to_write = []
        for item in shared_data:
            # 对导出excel的行进行调整，当prefix为空时，不填入narrator
            data_to_write.append([item["Prefix"], item["Original"], item["Translation"], "", item["Location"], item["Identifier"]])
        for row in tqdm(data_to_write, desc="写入 Excel"):
            sheet.append(row)

        # 设置列宽
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 25
        workbook.save(output_excel_file)


if __name__ == '__main__':
    start_time = time.time()
    game_root = "."
    tl_folder = os.path.join(game_root, "game", "tl")

    language_folder = input("请输入要操作的语言文件夹名称 (例如: chinese): ")
    language_path = os.path.join(tl_folder, language_folder)
    if not os.path.isdir(language_path):
        print("指定的语言文件夹不存在！")
        exit()

    excel_file = os.path.join(game_root, f"{language_folder}.xlsx")

    if os.path.exists(excel_file):
        print(f"文件 {excel_file} 已存在，将会被覆盖。")
    export_to_excel(language_path, language_folder, excel_file)
    print(f"翻译数据已导出到 {excel_file}")

    end_time = time.time()
    print(f"总耗时：{end_time - start_time} 秒")